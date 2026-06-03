import csv
import io
import json

from conftest import create_bound_models
from fastapi.testclient import TestClient
from openpyxl import Workbook, load_workbook

from app.db.models import FAQEntry


def _create_faq_kb(client: TestClient) -> str:
    chat_id, embedding_id = create_bound_models(client)
    response = client.post(
        "/api/v1/knowledge-bases",
        json={
            "name": "FAQ import KB",
            "kb_type": "faq",
            "summary_model_id": chat_id,
            "embedding_model_id": embedding_id,
        },
    )
    assert response.status_code == 201, response.text
    return response.json()["id"]


def _csv_bytes(rows: list[dict[str, str]]) -> bytes:
    buffer = io.StringIO()
    writer = csv.DictWriter(buffer, fieldnames=["question", "answer", "metadata", "enabled", "tag_id"])
    writer.writeheader()
    writer.writerows(rows)
    return buffer.getvalue().encode("utf-8")


def _xlsx_bytes(rows: list[dict[str, str]]) -> bytes:
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(["question", "answer", "metadata", "enabled", "tag_id"])
    for row in rows:
        sheet.append(
            [row.get("question"), row.get("answer"), row.get("metadata"), row.get("enabled"), row.get("tag_id")]
        )
    buffer = io.BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


def test_faq_csv_import_append_reports_failures_and_indexes_successes(client: TestClient, fake_vector_store):
    kb_id = _create_faq_kb(client)
    tag = client.post(f"/api/v1/knowledge-bases/{kb_id}/tags", json={"name": "导入标签"}).json()
    payload = _csv_bytes(
        [
            {
                "question": "如何申请退款？",
                "answer": "在订单页面提交退款申请。",
                "metadata": json.dumps({"source": "csv"}, ensure_ascii=False),
                "enabled": "true",
                "tag_id": tag["id"],
            },
            {"question": "缺少答案", "answer": "", "metadata": "{}", "enabled": "true", "tag_id": ""},
        ]
    )

    response = client.post(
        f"/api/v1/knowledge-bases/{kb_id}/faqs/import",
        data={"mode": "append"},
        files={"file": ("faqs.csv", payload, "text/csv")},
    )

    assert response.status_code == 200, response.text
    result = response.json()
    assert result["imported"] == 1
    assert result["failed"] == 1
    assert result["task_id"]
    assert result["processed"] == 2
    assert result["succeeded"] == 1
    assert result["status"] == "completed"
    assert result["errors"][0]["row"] == 3
    progress = client.get(f"/api/v1/knowledge-bases/{kb_id}/faqs/import-progress/{result['task_id']}")
    assert progress.status_code == 200, progress.text
    assert progress.json()["processed"] == 2
    last_result = client.get(f"/api/v1/knowledge-bases/{kb_id}/faqs/import-last-result")
    assert last_result.status_code == 200, last_result.text
    assert last_result.json()["task_id"] == result["task_id"]
    assert last_result.json()["display_status"] == "open"
    close_response = client.put(
        f"/api/v1/knowledge-bases/{kb_id}/faqs/import-last-result/display-status",
        json={"display_status": "close"},
    )
    assert close_response.status_code == 200, close_response.text
    assert close_response.json()["display_status"] == "close"
    faqs = client.get(f"/api/v1/knowledge-bases/{kb_id}/faqs").json()
    assert faqs[0]["question"] == "如何申请退款？"
    assert faqs[0]["tag_id"] == tag["id"]
    assert fake_vector_store.points[-1]["payload"]["metadata"]["source"] == "csv"
    assert fake_vector_store.points[-1]["payload"]["tag_id"] == tag["id"]


def test_faq_xlsx_import_replace_soft_deletes_existing_entries(client: TestClient, db_session):
    kb_id = _create_faq_kb(client)
    existing = client.post(
        f"/api/v1/knowledge-bases/{kb_id}/faqs",
        json={"question": "旧问题", "answer": "旧答案"},
    ).json()
    payload = _xlsx_bytes(
        [
            {"question": "新问题 A", "answer": "新答案 A", "metadata": "{}", "enabled": "true", "tag_id": ""},
            {"question": "新问题 B", "answer": "新答案 B", "metadata": "{}", "enabled": "false", "tag_id": ""},
        ]
    )

    response = client.post(
        f"/api/v1/knowledge-bases/{kb_id}/faqs/import",
        data={"mode": "replace"},
        files={"file": ("faqs.xlsx", payload, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
    )

    assert response.status_code == 200, response.text
    assert response.json()["imported"] == 2
    assert db_session.get(FAQEntry, existing["id"]).deleted_at is not None
    assert [item["question"] for item in client.get(f"/api/v1/knowledge-bases/{kb_id}/faqs").json()] == [
        "新问题 B",
        "新问题 A",
    ]


def test_faq_export_supports_csv_and_xlsx(client: TestClient):
    kb_id = _create_faq_kb(client)
    client.post(
        f"/api/v1/knowledge-bases/{kb_id}/faqs",
        json={"question": "导出问题", "answer": "导出答案", "metadata": {"source": "manual"}, "enabled": True},
    )

    csv_response = client.get(f"/api/v1/knowledge-bases/{kb_id}/faqs/export", params={"format": "csv"})
    assert csv_response.status_code == 200
    assert "text/csv" in csv_response.headers["content-type"]
    assert "question,similar_questions,answer,metadata,enabled,tag_id" in csv_response.text
    assert "导出问题" in csv_response.text

    xlsx_response = client.get(f"/api/v1/knowledge-bases/{kb_id}/faqs/export", params={"format": "xlsx"})
    assert xlsx_response.status_code == 200
    workbook = load_workbook(io.BytesIO(xlsx_response.content))
    sheet = workbook.active
    assert [cell.value for cell in sheet[1]] == [
        "question",
        "similar_questions",
        "answer",
        "metadata",
        "enabled",
        "tag_id",
    ]
    assert sheet["A2"].value == "导出问题"
