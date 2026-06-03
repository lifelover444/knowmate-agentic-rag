import csv
import io
import json
import time
from dataclasses import dataclass
from datetime import UTC, datetime

from openpyxl import Workbook, load_workbook
from sqlalchemy import select
from sqlalchemy.orm import Session

from app.core.config import Settings
from app.db.models import FAQImportResult
from app.db.repositories.faq import FAQEntryRepository
from app.db.repositories.tag import KnowledgeTagRepository
from app.schemas.faq import FAQEntryCreate
from app.services.faq import FAQEntryService

FAQ_COLUMNS = ["question", "similar_questions", "answer", "metadata", "enabled", "tag_id"]


@dataclass
class FAQImportError:
    row: int
    error: str
    question: str | None = None


class FAQImportExportService:
    def __init__(self, db: Session, settings: Settings, vector_store, embedder=None) -> None:
        self.db = db
        self.settings = settings
        self.vector_store = vector_store
        self.embedder = embedder
        self.faqs = FAQEntryRepository(db)

    def import_file(self, *, kb_id: str, filename: str, data: bytes, mode: str) -> dict:
        if mode not in {"append", "replace"}:
            raise ValueError("导入模式必须是 append 或 replace")
        started_at = time.perf_counter()
        rows = _read_rows(filename, data)
        result = self._create_import_result(kb_id=kb_id, mode=mode, total=len(rows))
        service = FAQEntryService(self.db, self.settings, self.vector_store, embedder=self.embedder)
        if mode == "replace":
            for entry in self.faqs.list_by_knowledge_base(kb_id):
                service.delete(entry)

        imported = 0
        errors: list[FAQImportError] = []
        for row_number, row in rows:
            try:
                payload = self._payload(kb_id, row)
                service.create(kb_id, payload)
                imported += 1
            except Exception as exc:
                errors.append(FAQImportError(row=row_number, question=(row.get("question") or None), error=str(exc)))
        result.status = "completed"
        result.progress = 100
        result.processed = len(rows)
        result.succeeded = imported
        result.failed = len(errors)
        result.failures_json = [error.__dict__ for error in errors]
        result.error_summary = "；".join(f"第 {error.row} 行：{error.error}" for error in errors[:5]) or None
        result.processing_time_ms = max(0, int((time.perf_counter() - started_at) * 1000))
        result.imported_at = datetime.now(UTC)
        result.updated_at = datetime.now(UTC)
        self.db.add(result)
        self.db.commit()
        self.db.refresh(result)
        return self._result_payload(result) | {
            "imported": imported,
            "failed": len(errors),
            "errors": [error.__dict__ for error in errors],
        }

    def get_import_progress(self, *, kb_id: str, task_id: str) -> dict | None:
        result = self.db.scalar(
            select(FAQImportResult).where(
                FAQImportResult.task_id == task_id,
                FAQImportResult.tenant_id == self.settings.default_tenant_id,
                FAQImportResult.knowledge_base_id == kb_id,
            )
        )
        return self._result_payload(result) if result is not None else None

    def get_last_import_result(self, *, kb_id: str) -> dict | None:
        result = self.db.scalar(
            select(FAQImportResult)
            .where(
                FAQImportResult.tenant_id == self.settings.default_tenant_id,
                FAQImportResult.knowledge_base_id == kb_id,
            )
            .order_by(FAQImportResult.created_at.desc(), FAQImportResult.task_id.desc())
            .limit(1)
        )
        return self._result_payload(result) if result is not None else None

    def update_last_import_display_status(self, *, kb_id: str, display_status: str) -> dict:
        if display_status not in {"open", "close"}:
            raise ValueError("display_status 必须是 open 或 close")
        result = self.db.scalar(
            select(FAQImportResult)
            .where(
                FAQImportResult.tenant_id == self.settings.default_tenant_id,
                FAQImportResult.knowledge_base_id == kb_id,
            )
            .order_by(FAQImportResult.created_at.desc(), FAQImportResult.task_id.desc())
            .limit(1)
        )
        if result is None:
            raise LookupError("FAQ 导入结果不存在")
        result.display_status = display_status
        result.updated_at = datetime.now(UTC)
        self.db.add(result)
        self.db.commit()
        self.db.refresh(result)
        return self._result_payload(result)

    def export_rows(self, kb_id: str) -> list[dict[str, str]]:
        rows: list[dict[str, str]] = []
        for entry in self.faqs.list_by_knowledge_base(kb_id):
            rows.append(
                {
                    "question": entry.question,
                    "similar_questions": "##".join(entry.similar_questions or []),
                    "answer": entry.answer,
                    "metadata": json.dumps(entry.faq_metadata or {}, ensure_ascii=False),
                    "enabled": "true" if entry.enabled else "false",
                    "tag_id": entry.tag_id or "",
                }
            )
        return rows

    def export_csv(self, kb_id: str) -> bytes:
        buffer = io.StringIO()
        writer = csv.DictWriter(buffer, fieldnames=FAQ_COLUMNS, lineterminator="\n")
        writer.writeheader()
        writer.writerows(self.export_rows(kb_id))
        return buffer.getvalue().encode("utf-8-sig")

    def export_xlsx(self, kb_id: str) -> bytes:
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "FAQs"
        sheet.append(FAQ_COLUMNS)
        for row in self.export_rows(kb_id):
            sheet.append([row[column] for column in FAQ_COLUMNS])
        buffer = io.BytesIO()
        workbook.save(buffer)
        return buffer.getvalue()

    def _payload(self, kb_id: str, row: dict[str, str]) -> FAQEntryCreate:
        question = (row.get("question") or "").strip()
        answer = (row.get("answer") or "").strip()
        if not question:
            raise ValueError("问题不能为空")
        if not answer:
            raise ValueError("答案不能为空")
        metadata_text = (row.get("metadata") or "{}").strip() or "{}"
        try:
            metadata = json.loads(metadata_text)
        except json.JSONDecodeError as exc:
            raise ValueError("metadata 必须是合法 JSON") from exc
        if not isinstance(metadata, dict):
            raise ValueError("metadata 必须是 JSON object")
        tag_id = (row.get("tag_id") or "").strip() or None
        if tag_id is not None:
            tag = KnowledgeTagRepository(self.db).get(tag_id, self.settings.default_tenant_id)
            if tag is None or tag.knowledge_base_id != kb_id:
                raise ValueError("标签不存在")
        return FAQEntryCreate(
            question=question,
            similar_questions=_parse_list(row.get("similar_questions")),
            answer=answer,
            metadata=metadata,
            enabled=_parse_bool(row.get("enabled")),
            tag_id=tag_id,
        )

    def _create_import_result(self, *, kb_id: str, mode: str, total: int) -> FAQImportResult:
        now = datetime.now(UTC)
        result = FAQImportResult(
            tenant_id=self.settings.default_tenant_id,
            knowledge_base_id=kb_id,
            status="processing",
            progress=0 if total else 100,
            total=total,
            processed=0,
            succeeded=0,
            failed=0,
            failures_json=[],
            import_mode=mode,
            display_status="open",
            processing_time_ms=0,
            created_at=now,
            updated_at=now,
        )
        self.db.add(result)
        self.db.commit()
        self.db.refresh(result)
        return result

    def _result_payload(self, result: FAQImportResult) -> dict:
        failures = result.failures_json or []
        return {
            "task_id": result.task_id,
            "knowledge_base_id": result.knowledge_base_id,
            "status": result.status,
            "progress": result.progress,
            "total": result.total,
            "processed": result.processed,
            "succeeded": result.succeeded,
            "imported": result.succeeded,
            "failed": result.failed,
            "failures": failures,
            "errors": failures,
            "error_summary": result.error_summary,
            "mode": result.import_mode,
            "import_mode": result.import_mode,
            "display_status": result.display_status,
            "processing_time_ms": result.processing_time_ms,
            "imported_at": result.imported_at,
        }


def _read_rows(filename: str, data: bytes) -> list[tuple[int, dict[str, str]]]:
    suffix = filename.lower().rsplit(".", 1)[-1] if "." in filename else ""
    if suffix == "csv":
        text = data.decode("utf-8-sig")
        reader = csv.DictReader(io.StringIO(text))
        return [(index, {key: value or "" for key, value in row.items()}) for index, row in enumerate(reader, start=2)]
    if suffix == "xlsx":
        workbook = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
        sheet = workbook.active
        rows = list(sheet.iter_rows(values_only=True))
        if not rows:
            return []
        headers = [str(value or "").strip() for value in rows[0]]
        result: list[tuple[int, dict[str, str]]] = []
        for index, values in enumerate(rows[1:], start=2):
            result.append((index, {header: str(value or "") for header, value in zip(headers, values, strict=False)}))
        return result
    raise ValueError("仅支持 CSV 或 XLSX 文件")


def _parse_bool(value: str | None) -> bool:
    normalized = (value or "true").strip().lower()
    return normalized not in {"false", "0", "no", "停用", "禁用"}


def _parse_list(value: str | None) -> list[str]:
    items: list[str] = []
    seen: set[str] = set()
    for raw in (value or "").split("##"):
        item = raw.strip()
        if not item or item in seen:
            continue
        seen.add(item)
        items.append(item)
    return items
